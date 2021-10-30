import openpyxl

def  cre_ex(list_jap,list_num1,list_num2):
    wb = openpyxl.Workbook()
    ws = wb.active

    # シートを取得
    sheet = wb['Sheet']
    # セルへ書き込む

    k = 0
    l = 0
    for i in range(0, len(list_jap)):
        sheet.cell(row = i + 4, column = 4).value = list_jap[i]
    for j in range(0, len(list_num1)):
            sheet.cell(row = k + 4, column = 5).value = list_num1[j]
            k = k + 1
    for k in range(len(list_num2)):
            sheet.cell(row = l + 4, column = 6).value = list_num2[k]
            l = l + 1

    sheet.column_dimensions['D'].width = 26
    # 保存する
    wb.save(r'C:\Users\taino\seculities_report\seculities\static\excel\Sample1.xlsx')
